import configparser
import logging
import time
import json
import threading
import os
import shutil
import tempfile
import zipfile
import datetime as dt
import hashlib
import re
import statistics
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
from flask import Blueprint, request, jsonify, g, current_app, make_response, session, redirect
from dotenv import load_dotenv
try:
    from prometheus_client import Counter, Histogram, generate_latest
    from prometheus_client import CONTENT_TYPE_LATEST
except Exception:  # pragma: no cover - optional dependency
    Counter = None
    Histogram = None
    generate_latest = None
    CONTENT_TYPE_LATEST = "text/plain; version=0.0.4"

from denpyo_toroku.app.util.response import Response
from denpyo_toroku.config import AppConfig

api_blueprint = Blueprint("api_blueprint", __name__)

if Counter is not None and Histogram is not None:
    REQUEST_COUNT = Counter(
        "denpyo_toroku_http_requests_total",
        "Total HTTP requests handled by Denpyo Toroku API",
        ["method", "endpoint", "status_code"],
    )
    REQUEST_LATENCY_SECONDS = Histogram(
        "denpyo_toroku_http_request_duration_seconds",
        "HTTP request latency in seconds for Denpyo Toroku API",
        ["method", "endpoint"],
    )
else:
    REQUEST_COUNT = None
    REQUEST_LATENCY_SECONDS = None

# 学習データのインメモリストア（リクエスト間で共有）
_training_data_store = []

# 学習データファイルパス（固定ファイル名）
_TRAINING_DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", "data", "training_data.xlsx")

# 学習状態（モジュール変数としてリクエスト間で共有）
_training_state = {
    'status': 'idle',       # idle | running | completed | failed
    'progress': '',
    'results': None,
    'error': None,
    'started_at': None,
    'finished_at': None,
    'previous_model_summary': None
}
_training_state_lock = threading.Lock()

_default_auth_username = os.environ.get("DENPYO_TOROKU_LOGIN_USERNAME", "admin")
_default_auth_password = os.environ.get("DENPYO_TOROKU_LOGIN_PASSWORD", "admin")
_OCI_MASKED_KEY = "[CONFIGURED]"
_DB_MASKED_SECRET = "[CONFIGURED]"
_DB_CONN_ENV_KEY = "ORACLE_26AI_CONNECTION_STRING"
_DB_ADB_OCID_ENV_KEY = "ADB_OCID"
_DB_REQUIRED_WALLET_FILES = ("cwallet.sso", "ewallet.pem", "sqlnet.ora", "tnsnames.ora")
_DB_UNNECESSARY_WALLET_FILES = ("README", "keystore.jks", "truststore.jks", "ojdbc.properties", "ewallet.p12")
_DB_TEST_TIMEOUT_SECONDS = 15  # DB接続テストの最大待機時間（秒）
_DB_TEST_EXECUTOR = ThreadPoolExecutor(max_workers=2, thread_name_prefix="db_test_")
_OCI_KEY_PATTERN = re.compile(
    r"-----BEGIN[\s\S]*?PRIVATE KEY-----[\s\S]*?-----END[\s\S]*?PRIVATE KEY-----"
)
_DEFAULT_OCI_REGION = "us-chicago-1"


def _to_bool(value, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "y", "on")
    return bool(value)


def _is_session_authenticated() -> bool:
    user = session.get("user", None)
    token = session.get("token", None)
    token_expiry_ts = session.get("token_expiry_ts", None)
    if not user or not token or not token_expiry_ts:
        return False
    return dt.datetime.now().timestamp() < float(token_expiry_ts)


def _normalize_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _expand_path(path_value: str, default: str) -> str:
    raw = _normalize_text(path_value, default)
    if not raw:
        raw = default
    return os.path.abspath(os.path.expanduser(raw))


def _extract_region_from_endpoint(service_endpoint: str) -> str:
    endpoint = _normalize_text(service_endpoint)
    if not endpoint:
        return _DEFAULT_OCI_REGION
    match = re.search(
        r"https?://inference\.generativeai\.([a-z0-9-]+)\.oci\.oraclecloud\.com",
        endpoint,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group(1)
    return _DEFAULT_OCI_REGION


def _project_root_path() -> Path:
    # denpyo_toroku/app/blueprints/api/api_blueprint.py -> リポジトリルート
    return Path(__file__).resolve().parents[4]


def _env_file_path() -> Path:
    return _project_root_path() / ".env"


def _runtime_oci_defaults() -> Dict[str, str]:
    config_path = _normalize_text(
        os.environ.get("OCI_CONFIG_PATH"),
        _normalize_text(getattr(AppConfig, "OCI_CONFIG_PATH", ""), "~/.oci/config"),
    )
    profile = _normalize_text(
        os.environ.get("OCI_CONFIG_PROFILE"),
        _normalize_text(getattr(AppConfig, "OCI_CONFIG_PROFILE", ""), "DEFAULT"),
    ) or "DEFAULT"
    service_endpoint = _normalize_text(
        os.environ.get("OCI_SERVICE_ENDPOINT"),
        _normalize_text(
            getattr(AppConfig, "OCI_SERVICE_ENDPOINT", ""),
            "https://inference.generativeai.us-chicago-1.oci.oraclecloud.com",
        ),
    )
    return {
        "config_path": config_path or "~/.oci/config",
        "profile": profile,
        "compartment_id": _normalize_text(
            os.environ.get("OCI_CONFIG_COMPARTMENT"),
            _normalize_text(getattr(AppConfig, "OCI_CONFIG_COMPARTMENT", "")),
        ),
        "service_endpoint": service_endpoint,
        "embedding_model_id": _normalize_text(
            os.environ.get("EMBEDDING_MODEL_ID"),
            _normalize_text(getattr(AppConfig, "EMBEDDING_MODEL_ID", ""), "cohere.embed-v4.0"),
        ),
    }


def _get_config_value(parser: configparser.ConfigParser, profile: str, key: str) -> str:
    if profile.upper() == "DEFAULT":
        return _normalize_text(parser.defaults().get(key))
    if parser.has_section(profile):
        return _normalize_text(parser.get(profile, key, fallback=""))
    fallback = _normalize_text(parser.defaults().get(key))
    if fallback:
        return fallback
    for section_name in parser.sections():
        value = _normalize_text(parser.get(section_name, key, fallback=""))
        if value:
            return value
    return ""


def _resolve_key_file_path(
    config_path: str,
    parser: configparser.ConfigParser,
    profile: str,
    explicit_key_file: str = "",
) -> str:
    config_dir = os.path.dirname(config_path) or "."
    configured_key = _normalize_text(explicit_key_file) or _get_config_value(parser, profile, "key_file")
    key_file = os.path.expanduser(configured_key) if configured_key else os.path.join(config_dir, "oci_api_key.pem")
    if not os.path.isabs(key_file):
        key_file = os.path.join(config_dir, key_file)
    return _expand_path(key_file, os.path.join(config_dir, "oci_api_key.pem"))


def _load_oci_settings_snapshot() -> Dict[str, Any]:
    defaults = _runtime_oci_defaults()
    config_path = _expand_path(defaults["config_path"], "~/.oci/config")
    profile = defaults["profile"] or "DEFAULT"
    parser = configparser.ConfigParser()
    if os.path.exists(config_path):
        parser.read(config_path)

    user_ocid = _get_config_value(parser, profile, "user")
    tenancy_ocid = _get_config_value(parser, profile, "tenancy")
    fingerprint = _get_config_value(parser, profile, "fingerprint")
    region = _get_config_value(parser, profile, "region") or _extract_region_from_endpoint(defaults["service_endpoint"])

    key_file = _resolve_key_file_path(config_path, parser, profile)
    key_content = ""
    if os.path.exists(key_file):
        try:
            with open(key_file, "r", encoding="utf-8") as key_reader:
                key_content = key_reader.read()
        except Exception:  # pragma: no cover - best effort read
            key_content = ""

    has_credentials = bool(
        user_ocid and tenancy_ocid and fingerprint and _normalize_text(key_content)
    )

    return {
        "settings": {
            "user_ocid": user_ocid,
            "tenancy_ocid": tenancy_ocid,
            "fingerprint": fingerprint,
            "region": region or _DEFAULT_OCI_REGION,
            "key_content": _OCI_MASKED_KEY if has_credentials else "",
            "config_path": config_path,
            "profile": profile,
            "key_file": key_file,
            "compartment_id": defaults["compartment_id"],
            "service_endpoint": defaults["service_endpoint"],
            "embedding_model_id": defaults["embedding_model_id"],
        },
        "has_credentials": has_credentials,
        "is_configured": has_credentials,
        "status": "configured" if has_credentials else "not_configured",
    }


def _validate_private_key_content(key_content: str) -> bool:
    content = _normalize_text(key_content)
    if not content:
        return False
    return bool(_OCI_KEY_PATTERN.search(content))


def _upsert_env_values(env_path: Path, updates: Dict[str, str]) -> None:
    lines: List[str] = []
    if env_path.exists():
        with open(env_path, "r", encoding="utf-8") as env_reader:
            lines = env_reader.read().splitlines()
    else:
        env_path.parent.mkdir(parents=True, exist_ok=True)

    replaced_keys = set()
    updated_lines: List[str] = []
    for line in lines:
        stripped = line.lstrip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            updated_lines.append(line)
            continue

        key = stripped.split("=", 1)[0].strip()
        if key in updates:
            updated_lines.append(f"{key}={updates[key]}")
            replaced_keys.add(key)
        else:
            updated_lines.append(line)

    for key, value in updates.items():
        if key not in replaced_keys:
            updated_lines.append(f"{key}={value}")

    with open(env_path, "w", encoding="utf-8") as env_writer:
        content = "\n".join(updated_lines).rstrip()
        env_writer.write(content + "\n" if content else "")


def _apply_runtime_oci_values(settings: Dict[str, str]) -> None:
    os.environ["OCI_CONFIG_PATH"] = settings["config_path"]
    os.environ["OCI_CONFIG_PROFILE"] = settings["profile"]
    os.environ["OCI_CONFIG_COMPARTMENT"] = settings["compartment_id"]
    os.environ["OCI_SERVICE_ENDPOINT"] = settings["service_endpoint"]
    os.environ["EMBEDDING_MODEL_ID"] = settings["embedding_model_id"]

    AppConfig.OCI_CONFIG_PATH = settings["config_path"]
    AppConfig.OCI_CONFIG_PROFILE = settings["profile"]
    AppConfig.OCI_CONFIG_COMPARTMENT = settings["compartment_id"]
    AppConfig.OCI_SERVICE_ENDPOINT = settings["service_endpoint"]
    AppConfig.EMBEDDING_MODEL_ID = settings["embedding_model_id"]


def _save_oci_settings(settings_payload: Dict[str, Any]) -> Dict[str, Any]:
    current_snapshot = _load_oci_settings_snapshot()
    current_settings = current_snapshot["settings"]

    config_path = _expand_path(
        _normalize_text(settings_payload.get("config_path"), current_settings.get("config_path", "~/.oci/config")),
        "~/.oci/config",
    )
    profile = _normalize_text(settings_payload.get("profile"), current_settings.get("profile", "DEFAULT")) or "DEFAULT"
    service_endpoint = _normalize_text(
        settings_payload.get("service_endpoint"),
        current_settings.get("service_endpoint", "https://inference.generativeai.us-chicago-1.oci.oraclecloud.com"),
    )

    parser = configparser.ConfigParser()
    if os.path.exists(config_path):
        parser.read(config_path)

    key_file = _resolve_key_file_path(
        config_path=config_path,
        parser=parser,
        profile=profile,
        explicit_key_file=_normalize_text(settings_payload.get("key_file")),
    )
    provided_key_content = settings_payload.get("key_content")
    provided_key_content_normalized = _normalize_text(provided_key_content)
    has_new_key = bool(
        provided_key_content_normalized and provided_key_content_normalized != _OCI_MASKED_KEY
    )

    if has_new_key:
        if not _validate_private_key_content(provided_key_content_normalized):
            raise ValueError("Invalid PEM format for private key.")
        key_dir = os.path.dirname(key_file) or "."
        os.makedirs(key_dir, exist_ok=True)
        try:
            os.chmod(key_dir, 0o700)
        except Exception:  # pragma: no cover - filesystem dependent
            pass
        with open(key_file, "w", encoding="utf-8") as key_writer:
            key_writer.write(str(provided_key_content))
        try:
            os.chmod(key_file, 0o600)
        except Exception:  # pragma: no cover - filesystem dependent
            pass
    elif not os.path.exists(key_file):
        raise ValueError("Private key is required for initial OCI configuration.")

    user_ocid = _normalize_text(settings_payload.get("user_ocid"), current_settings.get("user_ocid", ""))
    tenancy_ocid = _normalize_text(settings_payload.get("tenancy_ocid"), current_settings.get("tenancy_ocid", ""))
    fingerprint = _normalize_text(settings_payload.get("fingerprint"), current_settings.get("fingerprint", ""))
    region = _normalize_text(settings_payload.get("region"), current_settings.get("region", ""))
    if not region:
        region = _extract_region_from_endpoint(service_endpoint)

    if profile.upper() == "DEFAULT":
        target_section = parser["DEFAULT"]
    else:
        if not parser.has_section(profile):
            parser.add_section(profile)
        target_section = parser[profile]

    target_section["user"] = user_ocid
    target_section["tenancy"] = tenancy_ocid
    target_section["fingerprint"] = fingerprint
    target_section["region"] = region
    target_section["key_file"] = key_file

    config_dir = os.path.dirname(config_path) or "."
    os.makedirs(config_dir, exist_ok=True)
    try:
        os.chmod(config_dir, 0o700)
    except Exception:  # pragma: no cover - filesystem dependent
        pass

    with open(config_path, "w", encoding="utf-8") as config_writer:
        parser.write(config_writer)
    try:
        os.chmod(config_path, 0o600)
    except Exception:  # pragma: no cover - filesystem dependent
        pass

    settings_for_env = {
        "config_path": config_path,
        "profile": profile,
        "compartment_id": _normalize_text(
            settings_payload.get("compartment_id"),
            current_settings.get("compartment_id", ""),
        ),
        "service_endpoint": service_endpoint,
        "embedding_model_id": _normalize_text(
            settings_payload.get("embedding_model_id"),
            current_settings.get("embedding_model_id", "cohere.embed-v4.0"),
        )
        or "cohere.embed-v4.0",
    }

    _upsert_env_values(
        _env_file_path(),
        {
            "OCI_CONFIG_PATH": settings_for_env["config_path"],
            "OCI_CONFIG_PROFILE": settings_for_env["profile"],
            "OCI_CONFIG_COMPARTMENT": settings_for_env["compartment_id"],
            "OCI_SERVICE_ENDPOINT": settings_for_env["service_endpoint"],
            "EMBEDDING_MODEL_ID": settings_for_env["embedding_model_id"],
        },
    )
    load_dotenv(_env_file_path(), override=True)
    _apply_runtime_oci_values(settings_for_env)

    snapshot = _load_oci_settings_snapshot()
    snapshot["status"] = "saved"
    snapshot["is_configured"] = snapshot["has_credentials"]
    return snapshot


def _build_oci_test_config(request_settings: Dict[str, Any]) -> Dict[str, str]:
    snapshot = _load_oci_settings_snapshot()["settings"]
    parser = configparser.ConfigParser()

    config_path = _expand_path(
        _normalize_text(request_settings.get("config_path"), snapshot.get("config_path", "~/.oci/config")),
        "~/.oci/config",
    )
    if os.path.exists(config_path):
        parser.read(config_path)
    profile = _normalize_text(request_settings.get("profile"), snapshot.get("profile", "DEFAULT")) or "DEFAULT"
    key_file = _resolve_key_file_path(config_path, parser, profile)

    key_content = _normalize_text(request_settings.get("key_content"), snapshot.get("key_content", ""))
    if key_content == _OCI_MASKED_KEY or not key_content:
        if os.path.exists(key_file):
            with open(key_file, "r", encoding="utf-8") as key_reader:
                key_content = key_reader.read()
        else:
            key_content = ""

    service_endpoint = _normalize_text(request_settings.get("service_endpoint"), snapshot.get("service_endpoint", ""))
    region = _normalize_text(request_settings.get("region"), snapshot.get("region", ""))
    if not region:
        region = _extract_region_from_endpoint(service_endpoint)

    return {
        "user": _normalize_text(request_settings.get("user_ocid"), snapshot.get("user_ocid", "")),
        "tenancy": _normalize_text(request_settings.get("tenancy_ocid"), snapshot.get("tenancy_ocid", "")),
        "fingerprint": _normalize_text(request_settings.get("fingerprint"), snapshot.get("fingerprint", "")),
        "key_content": key_content,
        "region": region,
    }


def _parse_db_connection_string(connection_string: str) -> Dict[str, str]:
    value = _normalize_text(connection_string)
    if not value or "/" not in value or "@" not in value:
        return {"username": "", "password": "", "dsn": ""}
    user_pass, dsn = value.rsplit("@", 1)
    if "/" not in user_pass:
        return {"username": "", "password": "", "dsn": ""}
    username, password = user_pass.split("/", 1)
    return {
        "username": _normalize_text(username),
        "password": _normalize_text(password),
        "dsn": _normalize_text(dsn),
    }


def _build_db_connection_string(username: str, password: str, dsn: str) -> str:
    return f"{username}/{password}@{dsn}"


def _db_wallet_location(create_if_missing: bool = False) -> Optional[str]:
    tns_admin = _normalize_text(os.environ.get("TNS_ADMIN"))
    if tns_admin:
        path = _expand_path(tns_admin, tns_admin)
        if create_if_missing:
            os.makedirs(path, exist_ok=True)
            return path
        return path if os.path.exists(path) else None

    oracle_client_lib_dir = _normalize_text(os.environ.get("ORACLE_CLIENT_LIB_DIR"))
    if oracle_client_lib_dir:
        wallet_path = os.path.join(_expand_path(oracle_client_lib_dir, oracle_client_lib_dir), "network", "admin")
        # reference behavior: TNS_ADMIN is derived from ORACLE_CLIENT_LIB_DIR/network/admin
        os.environ["TNS_ADMIN"] = wallet_path
    else:
        wallet_path = os.path.join(str(_project_root_path()), "denpyo_toroku", "data", "wallet")

    if create_if_missing:
        os.makedirs(wallet_path, exist_ok=True)
        return wallet_path
    return wallet_path if os.path.exists(wallet_path) else None


def _extract_services_from_tnsnames(wallet_location: Optional[str]) -> List[str]:
    if not wallet_location:
        return []
    tnsnames_path = os.path.join(wallet_location, "tnsnames.ora")
    if not os.path.exists(tnsnames_path):
        return []
    try:
        with open(tnsnames_path, "r", encoding="utf-8") as reader:
            content = reader.read()
        matches = re.findall(r"^([A-Za-z0-9_-]+)\s*=", content, re.MULTILINE)
        return sorted(set(_normalize_text(item) for item in matches if _normalize_text(item)))
    except Exception:
        return []


def _wallet_is_ready(wallet_location: Optional[str]) -> bool:
    if not wallet_location:
        return False
    return all(os.path.exists(os.path.join(wallet_location, file_name)) for file_name in _DB_REQUIRED_WALLET_FILES)


def _load_database_settings_snapshot() -> Dict[str, Any]:
    conn_info = _parse_db_connection_string(os.environ.get(_DB_CONN_ENV_KEY, ""))
    adb_ocid = _normalize_text(os.environ.get(_DB_ADB_OCID_ENV_KEY, ""))
    wallet_location = _db_wallet_location(create_if_missing=False)
    wallet_uploaded = _wallet_is_ready(wallet_location)
    available_services = _extract_services_from_tnsnames(wallet_location)

    has_core_settings = bool(conn_info["username"] and conn_info["password"] and conn_info["dsn"])
    is_configured = has_core_settings and wallet_uploaded

    return {
        "settings": {
            "username": conn_info["username"],
            "password": _DB_MASKED_SECRET if conn_info["password"] else "",
            "dsn": conn_info["dsn"],
            "adb_ocid": adb_ocid,
            "wallet_uploaded": wallet_uploaded,
            "available_services": available_services,
        },
        "wallet_location": wallet_location if wallet_uploaded else None,
        "is_connected": False,
        "is_configured": is_configured,
        "status": "configured" if is_configured else "not_configured",
    }


def _save_database_settings(settings_payload: Dict[str, Any]) -> Dict[str, Any]:
    current_snapshot = _load_database_settings_snapshot()
    current_settings = current_snapshot["settings"]
    current_conn_info = _parse_db_connection_string(os.environ.get(_DB_CONN_ENV_KEY, ""))

    username = _normalize_text(settings_payload.get("username"), current_conn_info["username"] or current_settings.get("username", ""))
    dsn = _normalize_text(settings_payload.get("dsn"), current_conn_info["dsn"] or current_settings.get("dsn", ""))

    incoming_password = settings_payload.get("password")
    incoming_password_text = _normalize_text(incoming_password)
    if incoming_password_text == _DB_MASKED_SECRET:
        password = current_conn_info["password"]
    elif incoming_password_text:
        password = incoming_password_text
    else:
        password = current_conn_info["password"]

    if not username or not password or not dsn:
        raise ValueError("ユーザー名・パスワード・DSN は必須です。")
    adb_ocid = _normalize_text(settings_payload.get("adb_ocid"), os.environ.get(_DB_ADB_OCID_ENV_KEY, ""))

    conn_string = _build_db_connection_string(username, password, dsn)
    _upsert_env_values(
        _env_file_path(),
        {
            _DB_CONN_ENV_KEY: conn_string,
            _DB_ADB_OCID_ENV_KEY: adb_ocid,
        },
    )
    load_dotenv(_env_file_path(), override=True)

    os.environ[_DB_CONN_ENV_KEY] = conn_string
    os.environ[_DB_ADB_OCID_ENV_KEY] = adb_ocid
    AppConfig.ADB_OCID = adb_ocid

    snapshot = _load_database_settings_snapshot()
    snapshot["status"] = "saved"
    return snapshot


def _database_env_connection_info(include_password: bool = False) -> Dict[str, Any]:
    conn_info = _parse_db_connection_string(os.environ.get(_DB_CONN_ENV_KEY, ""))
    adb_ocid = _normalize_text(os.environ.get(_DB_ADB_OCID_ENV_KEY, ""))
    region = _normalize_text(os.environ.get("OCI_REGION"), _normalize_text(getattr(AppConfig, "OCI_REGION", ""), "ap-osaka-1"))
    wallet_location = _db_wallet_location(create_if_missing=False)
    wallet_exists = _wallet_is_ready(wallet_location)
    available_services = _extract_services_from_tnsnames(wallet_location)

    if not conn_info["username"] or not conn_info["dsn"]:
        return {
            "success": False,
            "message": f"{_DB_CONN_ENV_KEY} 環境変数が未設定、または形式が不正です。",
            "username": None,
            "password": None,
            "dsn": None,
            "adb_ocid": adb_ocid or None,
            "region": region or None,
            "wallet_exists": wallet_exists,
            "wallet_location": wallet_location if wallet_exists else None,
            "available_services": available_services,
        }

    return {
        "success": True,
        "message": "環境変数から接続情報を取得しました。",
        "username": conn_info["username"],
        "password": conn_info["password"] if include_password else (_DB_MASKED_SECRET if conn_info["password"] else ""),
        "dsn": conn_info["dsn"],
        "adb_ocid": adb_ocid or None,
        "region": region or None,
        "wallet_exists": wallet_exists,
        "wallet_location": wallet_location if wallet_exists else None,
        "available_services": available_services,
    }


def _map_db_connection_error(error_text: str) -> str:
    if "DPY-6005" in error_text or "DPY-6000" in error_text:
        return "接続エラー: データベースが停止している可能性があります。"
    if "ORA-01017" in error_text:
        return "接続エラー: ユーザー名またはパスワードが正しくありません。"
    if "ORA-12154" in error_text:
        return "接続エラー: DSN が見つかりません。Wallet と tnsnames.ora を確認してください。"
    if "ORA-12541" in error_text:
        return "接続エラー: データベースサーバーに接続できません。"
    if "DPY-4011" in error_text:
        return "接続エラー: Wallet または TNS 設定に問題があります。"
    return f"接続エラー: {error_text}"


def _test_database_connection(settings_payload: Dict[str, Any]) -> Dict[str, Any]:
    payload = settings_payload if isinstance(settings_payload, dict) else {}
    snapshot = _load_database_settings_snapshot()["settings"]
    env_info = _database_env_connection_info(include_password=True)

    username = _normalize_text(payload.get("username"), snapshot.get("username", ""))
    dsn = _normalize_text(payload.get("dsn"), snapshot.get("dsn", ""))
    incoming_password = _normalize_text(payload.get("password"))
    if incoming_password == _DB_MASKED_SECRET or not incoming_password:
        password = _normalize_text(env_info.get("password") if env_info.get("success") else "")
    else:
        password = incoming_password

    if not username or not password or not dsn:
        return {
            "success": False,
            "message": "接続情報が不完全です。ユーザー名・パスワード・DSN を入力してください。",
            "details": None,
        }

    wallet_location = _db_wallet_location(create_if_missing=False)
    if not _wallet_is_ready(wallet_location):
        return {
            "success": False,
            "message": f"Wallet ディレクトリが未設定、または必要ファイルが不足しています: {wallet_location}",
            "details": None,
        }

    try:
        import oracledb  # type: ignore
    except Exception:
        return {
            "success": False,
            "message": "oracledb モジュールが未インストールです。`pip install oracledb` を実行してください。",
            "details": None,
        }

    connection = None
    started_at = time.time()
    try:
        connection = oracledb.connect(
            user=username,
            password=password,
            dsn=dsn,
            config_dir=wallet_location,
            wallet_location=wallet_location,
            wallet_password=password,
            tcp_connect_timeout=10,
        )
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1 FROM DUAL")
            cursor.fetchone()
        elapsed = round(time.time() - started_at, 2)
        return {
            "success": True,
            "message": "データベース接続に成功しました。",
            "details": {
                "username": username,
                "dsn": dsn,
                "wallet_location": wallet_location,
                "elapsed_sec": str(elapsed),
            },
        }
    except Exception as e:
        return {
            "success": False,
            "message": _map_db_connection_error(str(e)),
            "details": None,
        }
    finally:
        if connection is not None:
            try:
                connection.close()
            except Exception:
                pass


def _upload_database_wallet(uploaded_file: Any) -> Dict[str, Any]:
    filename = _normalize_text(getattr(uploaded_file, "filename", ""))
    if not filename.lower().endswith(".zip"):
        raise ValueError("ZIPファイルのみ対応しています。")

    wallet_location = _db_wallet_location(create_if_missing=True)
    if not wallet_location:
        raise ValueError("Wallet の保存先を決定できませんでした。")

    with tempfile.NamedTemporaryFile(prefix="wallet_", suffix=".zip", delete=False) as temp_writer:
        temp_path = temp_writer.name
    try:
        uploaded_file.save(temp_path)

        if os.path.exists(wallet_location):
            backup_dir = f"{wallet_location}_backup_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            shutil.move(wallet_location, backup_dir)
        os.makedirs(wallet_location, exist_ok=True)

        with zipfile.ZipFile(temp_path, "r") as zip_ref:
            zip_ref.extractall(wallet_location)

        for file_name in _DB_UNNECESSARY_WALLET_FILES:
            file_path = os.path.join(wallet_location, file_name)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass

        missing = [f for f in _DB_REQUIRED_WALLET_FILES if not os.path.exists(os.path.join(wallet_location, f))]
        if missing:
            raise ValueError(
                "必要な Wallet ファイルが不足しています: %s" % ", ".join(missing)
            )

        services = _extract_services_from_tnsnames(wallet_location)
        return {
            "success": True,
            "message": "Walletをアップロードしました。",
            "wallet_location": wallet_location,
            "available_services": services,
        }
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass


def _sanitize_training_items(training_data: Any) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    """Normalize raw training data and return accepted items plus rejection stats."""
    accepted: List[Dict[str, str]] = []
    stats = {
        "invalid_item_count": 0,
        "missing_field_count": 0,
        "empty_text_count": 0,
        "empty_label_count": 0,
    }

    if not isinstance(training_data, list):
        return accepted, stats

    for item in training_data:
        if not isinstance(item, dict):
            stats["invalid_item_count"] += 1
            continue
        if "text" not in item or "label" not in item:
            stats["missing_field_count"] += 1
            continue

        text_val = str(item.get("text", "")).strip()
        label_val = str(item.get("label", "")).strip()
        if not text_val:
            stats["empty_text_count"] += 1
            continue
        if not label_val:
            stats["empty_label_count"] += 1
            continue

        accepted.append({"text": text_val, "label": label_val})

    return accepted, stats


def _ordered_class_distribution(labels: List[str]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for label in labels:
        counts[label] = counts.get(label, 0) + 1
    ordered_items = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    return {k: v for k, v in ordered_items}


def _compute_text_length_stats(texts: List[str]) -> Dict[str, float]:
    if not texts:
        return {
            "avg": 0.0,
            "median": 0.0,
            "p95": 0.0,
            "min": 0.0,
            "max": 0.0
        }

    lengths = sorted(len(t) for t in texts)
    p95_idx = min(len(lengths) - 1, int(round((len(lengths) - 1) * 0.95)))
    return {
        "avg": float(round(sum(lengths) / len(lengths), 2)),
        "median": float(round(statistics.median(lengths), 2)),
        "p95": float(lengths[p95_idx]),
        "min": float(lengths[0]),
        "max": float(lengths[-1])
    }


def _build_training_profile(training_data: List[Dict[str, str]]) -> Dict[str, Any]:
    """学習データの品質プロファイルと推奨設定を算出する。"""
    texts = [str(item["text"]).strip() for item in training_data]
    labels = [str(item["label"]).strip() for item in training_data]
    total_samples = len(texts)
    class_distribution = _ordered_class_distribution(labels)
    num_classes = len(class_distribution)
    min_class_count = min(class_distribution.values()) if class_distribution else 0
    max_class_count = max(class_distribution.values()) if class_distribution else 0
    imbalance_ratio = (max_class_count / min_class_count) if min_class_count > 0 else 0.0

    normalized_texts = [" ".join(t.lower().split()) for t in texts]
    unique_texts = len(set(normalized_texts))
    duplicate_count = max(0, total_samples - unique_texts)
    duplicate_ratio = (duplicate_count / total_samples) if total_samples > 0 else 0.0
    short_text_count = sum(1 for t in texts if len(t) < 4)
    long_text_count = sum(1 for t in texts if len(t) > 300)
    text_length_stats = _compute_text_length_stats(texts)

    issue_details: List[Dict[str, str]] = []
    recommendations: List[str] = []

    if total_samples < 20:
        issue_details.append({
            "level": "error",
            "message": "サンプル数が 20 件未満のため、モデル品質が不安定になる可能性があります。"
        })
        recommendations.append("ベースライン品質のため、総サンプル数を少なくとも 100 件まで増やしてください。")
    elif total_samples < 100:
        issue_details.append({
            "level": "warning",
            "message": "サンプル数が少ないため、再学習ごとに指標が変動しやすいです。"
        })
        recommendations.append("重要な意図ごとに少なくとも 50 件のサンプルを目安にしてください。")

    if num_classes < 2:
        issue_details.append({
            "level": "error",
            "message": "学習には少なくとも 2 つ以上のクラスが必要です。"
        })
    elif num_classes < 4:
        issue_details.append({
            "level": "info",
            "message": "クラス数が少ないため、過学習に注意して監視してください。"
        })

    rare_classes = [k for k, v in class_distribution.items() if v < 5]
    critical_classes = [k for k, v in class_distribution.items() if v < 2]
    if critical_classes:
        issue_details.append({
            "level": "error",
            "message": "2 件未満のサンプルしかないクラスがあります: %s" % ", ".join(critical_classes[:5])
        })
        recommendations.append("層化検証を有効にするため、2 件未満のクラスにサンプルを追加してください。")
    elif rare_classes:
        issue_details.append({
            "level": "warning",
            "message": "サンプルが少ないクラス（5 件未満）: %s" % ", ".join(rare_classes[:6])
        })
        recommendations.append("少数クラスの例を増やして再現率の改善を図ってください。")

    if imbalance_ratio >= 6:
        issue_details.append({
            "level": "warning",
            "message": "クラスの偏りが大きいです（%.1f:1）。" % imbalance_ratio
        })
        recommendations.append("クラスのリバランスを有効にし、少数クラスのデータを重点的に追加してください。")
    elif imbalance_ratio >= 3:
        issue_details.append({
            "level": "info",
            "message": "クラスの偏りが見られます（%.1f:1）。" % imbalance_ratio
        })
        recommendations.append("少数クラスに対してバランス・アップサンプルを検討してください。")

    if duplicate_ratio >= 0.2:
        issue_details.append({
            "level": "warning",
            "message": "テキスト重複率が高いです（%.1f%%）。" % (duplicate_ratio * 100)
        })
        recommendations.append("過学習を避けるため、重複テキストを削除してください。")
    elif duplicate_count > 0:
        issue_details.append({
            "level": "info",
            "message": "重複テキストを %d 件検出しました。" % duplicate_count
        })

    if short_text_count > 0:
        issue_details.append({
            "level": "info",
            "message": "短すぎるテキスト（4 文字未満）が %d 件あります（曖昧になりやすい）。" % short_text_count
        })
    if long_text_count > 0:
        issue_details.append({
            "level": "info",
            "message": "長すぎるテキスト（300 文字超）を %d 件検出しました。意図の例は簡潔にすることを検討してください。" % long_text_count
        })

    # ヒューリスティックな健全性スコア（0〜100）
    health_score = 100.0
    if total_samples < 100:
        health_score -= min(25.0, (100 - total_samples) * 0.25)
    if num_classes < 2:
        health_score -= 40.0
    if min_class_count < 5:
        health_score -= min(20.0, (5 - min_class_count) * 4.0)
    if imbalance_ratio > 3:
        health_score -= min(20.0, (imbalance_ratio - 3.0) * 4.0)
    health_score -= min(15.0, duplicate_ratio * 100.0 * 0.5)
    if short_text_count > total_samples * 0.15 and total_samples > 0:
        health_score -= 5.0
    health_score = max(0.0, min(100.0, health_score))
    health_score = float(round(health_score, 1))

    readiness = "high" if health_score >= 80 else ("medium" if health_score >= 60 else "low")
    quality_gate_passed = (
        total_samples >= 20 and
        num_classes >= 2 and
        min_class_count >= 2
    )

    # データセット診断に基づく推奨パラメータ
    suggested = {
        "test_size": 0.15,
        "n_estimators": 300,
        "learning_rate": 0.05,
        "max_depth": 6,
        "algorithm_strategy": "auto",
        "compare_baselines": True,
        "auto_tune": total_samples >= 120,
        "rebalance_strategy": "balanced_upsample" if imbalance_ratio >= 3 else "none"
    }
    if total_samples < 120:
        suggested["test_size"] = 0.2
        suggested["n_estimators"] = 200
        suggested["max_depth"] = 4
        suggested["learning_rate"] = 0.07
        suggested["auto_tune"] = False
    elif total_samples > 6000:
        suggested["test_size"] = 0.12
        suggested["n_estimators"] = 450
        suggested["max_depth"] = 7
        suggested["learning_rate"] = 0.04
        suggested["auto_tune"] = True

    if num_classes >= 20:
        suggested["max_depth"] = min(int(suggested["max_depth"]), 6)
        suggested["compare_baselines"] = True

    # API 制約に合わせてクランプ
    suggested["test_size"] = float(max(0.05, min(0.40, suggested["test_size"])))
    suggested["n_estimators"] = int(max(50, min(1000, suggested["n_estimators"])))
    suggested["learning_rate"] = float(max(0.01, min(0.30, suggested["learning_rate"])))
    suggested["max_depth"] = int(max(2, min(10, suggested["max_depth"])))

    if quality_gate_passed and not recommendations:
        recommendations.append("データ品質は良好です。学習を進め、結果の Macro-F1 を監視してください。")
    if not quality_gate_passed:
        recommendations.append("不安定なモデルを避けるため、学習開始前にブロッカーとなるデータ問題を解消してください。")

    return {
        "total_samples": total_samples,
        "num_classes": num_classes,
        "class_distribution": class_distribution,
        "class_distribution_percent": {
            k: float(round((v / total_samples) * 100, 2)) if total_samples > 0 else 0.0
            for k, v in class_distribution.items()
        },
        "min_class_count": min_class_count,
        "max_class_count": max_class_count,
        "imbalance_ratio": float(round(imbalance_ratio, 2)) if imbalance_ratio else 0.0,
        "duplicate_count": duplicate_count,
        "duplicate_ratio": float(round(duplicate_ratio, 4)),
        "short_text_count": short_text_count,
        "long_text_count": long_text_count,
        "text_length_stats": text_length_stats,
        "health_score": health_score,
        "readiness": readiness,
        "quality_gate_passed": quality_gate_passed,
        "issue_details": issue_details,
        "issues": [item["message"] for item in issue_details],
        "recommendations": recommendations,
        "suggested_params": suggested
    }


def _to_optional_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _extract_previous_model_summary(classifier: Any) -> Optional[Dict[str, Any]]:
    """読み込み済み分類器から、比較用の前回モデル要約を抽出する。"""
    if classifier is None or getattr(classifier, "classifier", None) is None:
        return None

    training_summary = getattr(classifier, "last_training_summary", None)
    if not isinstance(training_summary, dict):
        return None

    previous = {
        "algorithm": getattr(classifier, "algorithm_name", None),
        "model_timestamp": getattr(classifier, "model_timestamp", None),
        "test_accuracy": _to_optional_float(training_summary.get("test_accuracy")),
        "test_macro_f1": _to_optional_float(training_summary.get("test_macro_f1")),
        "test_weighted_f1": _to_optional_float(training_summary.get("test_weighted_f1")),
        "overfitting_gap": _to_optional_float(training_summary.get("overfitting_gap")),
        "selection_score": _to_optional_float(training_summary.get("selection_score")),
    }

    # 比較対象の指標がない場合は利用不可とする
    comparable_exists = any(
        previous[k] is not None
        for k in ("test_accuracy", "test_macro_f1", "overfitting_gap", "selection_score")
    )
    return previous if comparable_exists else None


def _build_model_comparison(previous: Optional[Dict[str, Any]], current: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Build metric deltas between current result and previous model summary."""
    if not previous:
        return None

    curr_acc = _to_optional_float(current.get("test_accuracy"))
    curr_macro_f1 = _to_optional_float(current.get("test_macro_f1"))
    curr_gap = _to_optional_float(current.get("overfitting_gap"))
    curr_score = _to_optional_float(current.get("selection_score"))

    prev_acc = _to_optional_float(previous.get("test_accuracy"))
    prev_macro_f1 = _to_optional_float(previous.get("test_macro_f1"))
    prev_gap = _to_optional_float(previous.get("overfitting_gap"))
    prev_score = _to_optional_float(previous.get("selection_score"))

    acc_delta = (curr_acc - prev_acc) if curr_acc is not None and prev_acc is not None else None
    macro_delta = (curr_macro_f1 - prev_macro_f1) if curr_macro_f1 is not None and prev_macro_f1 is not None else None
    gap_delta = (curr_gap - prev_gap) if curr_gap is not None and prev_gap is not None else None
    score_delta = (curr_score - prev_score) if curr_score is not None and prev_score is not None else None

    improvement_signals = 0
    regression_signals = 0
    if acc_delta is not None:
        if acc_delta > 0.003:
            improvement_signals += 1
        elif acc_delta < -0.003:
            regression_signals += 1
    if macro_delta is not None:
        if macro_delta > 0.003:
            improvement_signals += 2
        elif macro_delta < -0.003:
            regression_signals += 2
    if gap_delta is not None:
        if gap_delta < -0.01:
            improvement_signals += 1
        elif gap_delta > 0.01:
            regression_signals += 1
    if score_delta is not None:
        if score_delta > 0.003:
            improvement_signals += 2
        elif score_delta < -0.003:
            regression_signals += 2

    improved = improvement_signals > regression_signals
    if improvement_signals == regression_signals:
        improved = None

    if improved is True:
        summary = "Current model appears better than previous model."
    elif improved is False:
        summary = "Current model appears weaker than previous model."
    else:
        summary = "Current model is similar to previous model."

    return {
        "improved": improved,
        "summary": summary,
        "test_accuracy_delta": acc_delta,
        "test_macro_f1_delta": macro_delta,
        "overfitting_gap_delta": gap_delta,
        "selection_score_delta": score_delta
    }


@api_blueprint.before_request
def before_request():
    g.response = Response()
    g.request_start_time = time.time()


@api_blueprint.after_request
def after_request(response):
    endpoint = request.endpoint or "unknown"
    method = request.method
    status_code = str(response.status_code)
    duration = max(time.time() - g.get("request_start_time", time.time()), 0)

    if REQUEST_COUNT is not None and REQUEST_LATENCY_SECONDS is not None:
        REQUEST_COUNT.labels(method=method, endpoint=endpoint, status_code=status_code).inc()
        REQUEST_LATENCY_SECONDS.labels(method=method, endpoint=endpoint).observe(duration)
    return response


@api_blueprint.route("/v1/me", methods=["GET"])
def get_current_user():
    """ERP-compatible current user endpoint."""
    user = session.get("user", None)
    user_id = session.get("user_id", None)
    role = session.get("role", "USER")
    if _is_session_authenticated():
        return jsonify({
            "authenticated": True,
            "user": user,
            "user_id": user_id,
            "role": role
        })
    return jsonify({"authenticated": False})


@api_blueprint.route("/v1/loginValidation", methods=["POST"])
def login_validation():
    """ERP-compatible login validation endpoint."""
    try:
        if request.is_json:
            data = request.get_json() or {}
            username = (data.get("username") or "").strip()
            password = data.get("password", "")
        else:
            username = (request.form.get("username", "") or "").strip()
            password = request.form.get("password", "")

        if not username or not password:
            return jsonify({"success": False, "message": "ユーザー名とパスワードは必須です"}), 400

        if username == _default_auth_username and password == _default_auth_password:
            session["user"] = username
            session["user_id"] = "admin-user-id"
            session["role"] = "ADMIN"
            session["token"] = hashlib.sha256(
                f"{username}{dt.datetime.now()}".encode()
            ).hexdigest()
            expiry_time = dt.datetime.now() + dt.timedelta(minutes=30)
            session["token_expiry_ts"] = int(expiry_time.timestamp())

            if request.is_json:
                return jsonify({
                    "success": True,
                    "message": "サインインしました",
                    "user": username,
                    "role": "ADMIN"
                })
            return redirect("/studio/", code=303)

        if request.is_json:
            return jsonify({"success": False, "message": "ユーザー名またはパスワードが正しくありません"}), 401
        return redirect("/studio/login?error=invalid", code=303)
    except Exception as e:
        logging.error("ログインエラー: %s", str(e), exc_info=True)
        return jsonify({"success": False, "message": "ログイン処理中にエラーが発生しました"}), 500


@api_blueprint.route("/logout", methods=["GET", "POST"])
def logout():
    """ERP-compatible logout endpoint."""
    session.pop("user", None)
    session.pop("user_id", None)
    session.pop("role", None)
    session.pop("token", None)
    session.pop("token_expiry_ts", None)
    return redirect("/studio/login", code=303)


@api_blueprint.route("/api/v1/auth/me", methods=["GET"])
def auth_me_compat():
    """Backward-compatible endpoint."""
    if not _is_session_authenticated():
        g.response.add_error_message("認証が必要です")
        return jsonify(g.response.get_result()), 401
    g.response.set_data({
        "email": session.get("user", _default_auth_username),
        "authenticated": True
    })
    return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/auth/login", methods=["POST"])
def auth_login_compat():
    """Backward-compatible endpoint."""
    data = request.get_json(silent=True) or {}
    username = (data.get("email") or data.get("username") or "").strip()
    password = data.get("password") or ""
    if username == _default_auth_username and password == _default_auth_password:
        session["user"] = username
        session["user_id"] = "admin-user-id"
        session["role"] = "ADMIN"
        session["token"] = hashlib.sha256(
            f"{username}{dt.datetime.now()}".encode()
        ).hexdigest()
        expiry_time = dt.datetime.now() + dt.timedelta(minutes=30)
        session["token_expiry_ts"] = int(expiry_time.timestamp())
        g.response.set_data({"email": username, "authenticated": True})
        return jsonify(g.response.get_result())
    g.response.add_error_message("ユーザー名またはパスワードが正しくありません")
    return jsonify(g.response.get_result()), 401


@api_blueprint.route("/api/v1/auth/logout", methods=["POST"])
def auth_logout_compat():
    """Backward-compatible endpoint."""
    session.pop("user", None)
    session.pop("user_id", None)
    session.pop("role", None)
    session.pop("token", None)
    session.pop("token_expiry_ts", None)
    g.response.set_data({"authenticated": False})
    return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/oci/settings", methods=["GET"])
def get_oci_settings():
    """Get current OCI application settings."""
    try:
        snapshot = _load_oci_settings_snapshot()
        g.response.set_data(snapshot)
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("OCI 設定の読み込みエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/oci/settings", methods=["POST"])
def save_oci_settings():
    """Save OCI application settings."""
    try:
        body = request.get_json(silent=True) or {}
        settings_payload = body.get("settings") if isinstance(body.get("settings"), dict) else body
        if not isinstance(settings_payload, dict):
            g.response.add_error_message("リクエストボディが不正です。")
            return jsonify(g.response.get_result()), 422

        # 必須項目は semantic-doc-search の OCI キー認証の挙動に合わせる
        required_fields = {
            "user_ocid": "ユーザー OCID",
            "tenancy_ocid": "テナンシ OCID",
            "fingerprint": "フィンガープリント",
            "region": "リージョン",
        }
        missing_fields = [
            label
            for field, label in required_fields.items()
            if not _normalize_text(settings_payload.get(field))
        ]
        if missing_fields:
            g.response.add_error_message("必須項目が未入力です: %s" % "、".join(missing_fields))
            return jsonify(g.response.get_result()), 422

        snapshot = _save_oci_settings(settings_payload)
        g.response.set_data(snapshot)
        return jsonify(g.response.get_result())
    except ValueError as e:
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 422
    except Exception as e:
        logging.error("OCI 設定の保存エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/oci/test", methods=["POST"])
def test_oci_connection():
    """Test OCI authentication settings."""
    try:
        body = request.get_json(silent=True) or {}
        settings_payload = body.get("settings") if isinstance(body.get("settings"), dict) else body
        if not isinstance(settings_payload, dict):
            g.response.add_error_message("リクエストボディが不正です。")
            return jsonify(g.response.get_result()), 422

        test_config = _build_oci_test_config(settings_payload)
        if not all([test_config["user"], test_config["tenancy"], test_config["fingerprint"], test_config["key_content"], test_config["region"]]):
            g.response.add_error_message(
                "OCI 認証情報が不足しています。ユーザー OCID / テナンシ OCID / フィンガープリント / リージョン / 秘密鍵が必須です。"
            )
            return jsonify(g.response.get_result()), 422

        import oci  # local import to avoid import-time dependency in non-OCI flows

        identity_client = oci.identity.IdentityClient(test_config)
        user_data = identity_client.get_user(test_config["user"]).data
        g.response.set_data({
            "success": True,
            "message": "OCI 接続テストに成功しました。",
            "details": {
                "user_name": getattr(user_data, "name", ""),
                "user_ocid": getattr(user_data, "id", test_config["user"]),
                "tenancy_ocid": test_config["tenancy"],
                "region": test_config["region"],
            },
        })
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("OCI 接続テストエラー: %s", e, exc_info=True)
        g.response.set_data({
            "success": False,
            "message": "OCI 接続テストに失敗しました: %s" % str(e),
            "details": None,
        })
        return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/database/settings", methods=["GET"])
def get_database_settings():
    """Get current database settings without performing a connection test."""
    try:
        snapshot = _load_database_settings_snapshot()
        g.response.set_data(snapshot)
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("DB 設定の読み込みエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/database/settings/env", methods=["GET"])
def get_database_env_settings():
    """Read DB connection info from environment variables."""
    try:
        include_password = _to_bool(request.args.get("include_password"), default=False)
        env_info = _database_env_connection_info(include_password=include_password)
        g.response.set_data(env_info)
        if env_info.get("success", False):
            return jsonify(g.response.get_result())
        return jsonify(g.response.get_result()), 200
    except Exception as e:
        logging.error("DB 環境変数読み込みエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/database/settings", methods=["POST"])
def save_database_settings():
    """Save DB settings to .env."""
    try:
        body = request.get_json(silent=True) or {}
        settings_payload = body.get("settings") if isinstance(body.get("settings"), dict) else body
        if not isinstance(settings_payload, dict):
            g.response.add_error_message("リクエストボディが不正です。")
            return jsonify(g.response.get_result()), 422

        snapshot = _save_database_settings(settings_payload)
        g.response.set_data(snapshot)
        return jsonify(g.response.get_result())
    except ValueError as e:
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 422
    except Exception as e:
        logging.error("DB 設定の保存エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/database/settings/test", methods=["POST"])
def test_database_connection():
    """Test DB connection with current or provided settings.
    
    タイムアウト制御付きでDB接続テストを実行し、メインスレッドをブロックしないようにします。
    参考: no.1-semantic-doc-search プロジェクトの実装
    """
    try:
        body = request.get_json(silent=True) or {}
        settings_payload = body.get("settings") if isinstance(body.get("settings"), dict) else body
        
        # タイムアウト付きでDB接続テストを実行（非ブロッキング）
        started_at = time.time()
        try:
            future = _DB_TEST_EXECUTOR.submit(_test_database_connection, settings_payload)
            result = future.result(timeout=_DB_TEST_TIMEOUT_SECONDS)
        except FuturesTimeoutError:
            elapsed = round(time.time() - started_at, 2)
            logging.warning("DB 接続テストがタイムアウトしました (%s秒)", elapsed)
            result = {
                "success": False,
                "message": f"接続テストがタイムアウトしました（{elapsed}秒）。データベースが起動しているか確認してください。",
                "details": {"timeout_seconds": str(_DB_TEST_TIMEOUT_SECONDS), "elapsed_sec": str(elapsed)},
            }
        except Exception as e:
            logging.error("DB 接続テスト実行エラー: %s", e, exc_info=True)
            result = {
                "success": False,
                "message": f"接続テスト実行エラー: {str(e)}",
                "details": None,
            }
        
        g.response.set_data(result)
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("DB 接続テストエラー: %s", e, exc_info=True)
        g.response.set_data({
            "success": False,
            "message": f"接続テストエラー: {str(e)}",
            "details": None,
        })
        return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/database/settings/wallet", methods=["POST"])
def upload_database_wallet():
    """Upload wallet zip and expand it to wallet directory."""
    try:
        if "file" not in request.files:
            g.response.add_error_message("アップロードファイルが見つかりません。")
            return jsonify(g.response.get_result()), 422
        uploaded_file = request.files.get("file")
        if uploaded_file is None or not uploaded_file.filename:
            g.response.add_error_message("ファイル名が不正です。")
            return jsonify(g.response.get_result()), 422

        result = _upload_database_wallet(uploaded_file)
        g.response.set_data(result)
        return jsonify(g.response.get_result())
    except ValueError as e:
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 422
    except Exception as e:
        logging.error("Wallet アップロードエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/version", methods=["GET"])
def get_version():
    """Version endpoint."""
    g.response.set_data({
        "service": "Intent Classifier Service",
        "version": "1.0.0",
        "status": "running",
        "endpoints": {
            "predict": "/api/v1/predict",
            "predict_single": "/api/v1/predict/single",
            "health": "/api/v1/health",
            "stats": "/api/v1/stats",
            "oci_settings_get": "/api/v1/oci/settings",
            "oci_settings_save": "/api/v1/oci/settings",
            "oci_test": "/api/v1/oci/test",
            "db_settings_get": "/api/v1/database/settings",
            "db_settings_save": "/api/v1/database/settings",
            "db_settings_env": "/api/v1/database/settings/env",
            "db_settings_test": "/api/v1/database/settings/test",
            "db_wallet_upload": "/api/v1/database/settings/wallet",
            "model_info": "/api/v1/model/info",
            "model_reload": "/api/v1/model/reload",
            "cache_clear": "/api/v1/cache/clear",
            "train": "/api/v1/train",
            "train_validate": "/api/v1/train/validate",
            "train_profile": "/api/v1/train/profile",
            "train_status": "/api/v1/train/status",
            "docs": "/"
        }
    })
    return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/health", methods=["GET"])
def health_check():
    """Health check endpoint."""
    classifier = current_app.classifier
    if classifier is None:
        # Return basic health even without classifier
        g.response.set_data({
            "status": "degraded",
            "message": "分類器が初期化されていません",
            "model_loaded": False,
            "version": "1.0.0"
        })
        return jsonify(g.response.get_result())

    try:
        health = classifier.health_check()
        g.response.set_data(health)
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("ヘルスチェックエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/metrics", methods=["GET"])
def get_metrics():
    """Prometheus metrics endpoint."""
    if generate_latest is None:
        return jsonify({"error": "prometheus_client がインストールされていません"}), 503

    payload = generate_latest()
    response = make_response(payload, 200)
    response.headers["Content-Type"] = CONTENT_TYPE_LATEST
    return response


@api_blueprint.route("/api/v1/stats", methods=["GET"])
def get_stats():
    """Statistics endpoint."""
    classifier = current_app.classifier
    if classifier is None:
        g.response.set_data({
            "performance": {
                "total_predictions": 0,
                "total_errors": 0,
                "error_rate": 0,
                "avg_prediction_time": 0,
                "p95_prediction_time": 0,
                "p99_prediction_time": 0,
                "min_prediction_time": 0,
                "max_prediction_time": 0
            },
            "cache": {
                "hits": 0,
                "misses": 0,
                "hit_rate": 0,
                "cache_size": 0,
                "max_size": 0
            },
            "model": None,
            "message": "分類器が初期化されていません"
        })
        return jsonify(g.response.get_result())

    try:
        stats = classifier.get_stats()
        g.response.set_data(stats)
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("統計取得エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/predict", methods=["POST"])
def predict():
    """Batch prediction endpoint."""
    classifier = current_app.classifier
    if classifier is None:
        g.response.add_error_message("分類器が初期化されていません")
        return jsonify(g.response.get_result()), 503

    start_time = dt.datetime.now()

    try:
        data = request.get_json()
        if not data or 'texts' not in data:
            g.response.add_error_message("「texts」フィールドは必須です")
            return jsonify(g.response.get_result()), 422

        texts = data['texts']
        if not isinstance(texts, list) or len(texts) == 0:
            g.response.add_error_message("「texts」は空でない配列である必要があります")
            return jsonify(g.response.get_result()), 422

        if len(texts) > 1000:
            g.response.add_error_message("1 回のリクエストで送信できるテキストは最大 1000 件です")
            return jsonify(g.response.get_result()), 422

        return_proba = data.get('return_proba', True)
        confidence_threshold = data.get('confidence_threshold', 0.5)
        top_k = data.get('top_k', 3)
        unknown_on_low_conf = _to_bool(data.get('unknown_on_low_conf', True), default=True)
        unknown_intent_label = data.get('unknown_intent_label', 'UNKNOWN')
        if not (0.0 <= float(confidence_threshold) <= 1.0):
            g.response.add_error_message("「confidence_threshold」は 0.0〜1.0 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422
        if not (1 <= int(top_k) <= 10):
            g.response.add_error_message("「top_k」は 1〜10 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422

        results = classifier.predict(
            texts=texts,
            return_proba=return_proba,
            confidence_threshold=float(confidence_threshold),
            top_k=int(top_k),
            unknown_on_low_conf=unknown_on_low_conf,
            unknown_intent_label=str(unknown_intent_label)
        )

        processing_time = (dt.datetime.now() - start_time).total_seconds()

        g.response.set_data({
            "results": results,
            "total": len(results),
            "processing_time": processing_time
        })
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("予測エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/predict/single", methods=["POST"])
def predict_single():
    """Single text prediction endpoint."""
    classifier = current_app.classifier
    if classifier is None:
        g.response.add_error_message("分類器が初期化されていません")
        return jsonify(g.response.get_result()), 503

    try:
        text = request.args.get('text', '')
        return_proba = request.args.get('return_proba', 'true').lower() == 'true'
        confidence_threshold = request.args.get('confidence_threshold', type=float)
        top_k = request.args.get('top_k', default=3, type=int)
        unknown_on_low_conf = request.args.get('unknown_on_low_conf', default='true').lower() == 'true'
        unknown_intent_label = request.args.get('unknown_intent_label', default='UNKNOWN', type=str)

        if not text:
            # JSON ボディも確認
            data = request.get_json(silent=True)
            if data:
                text = data.get('text', '')
                return_proba = data.get('return_proba', True)
                if 'confidence_threshold' in data:
                    confidence_threshold = data.get('confidence_threshold')
                if 'top_k' in data:
                    top_k = data.get('top_k')
                if 'unknown_on_low_conf' in data:
                    unknown_on_low_conf = _to_bool(data.get('unknown_on_low_conf'), default=True)
                if 'unknown_intent_label' in data:
                    unknown_intent_label = data.get('unknown_intent_label')

        if not text:
            g.response.add_error_message("「text」は必須です")
            return jsonify(g.response.get_result()), 422

        if confidence_threshold is None:
            confidence_threshold = 0.5
        if not (0.0 <= float(confidence_threshold) <= 1.0):
            g.response.add_error_message("「confidence_threshold」は 0.0〜1.0 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422
        if not (1 <= int(top_k) <= 10):
            g.response.add_error_message("「top_k」は 1〜10 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422

        results = classifier.predict(
            texts=[text],
            return_proba=return_proba,
            confidence_threshold=float(confidence_threshold),
            top_k=int(top_k),
            unknown_on_low_conf=_to_bool(unknown_on_low_conf, default=True),
            unknown_intent_label=str(unknown_intent_label)
        )

        g.response.set_data(results[0])
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("予測エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/cache/clear", methods=["POST"])
def clear_cache():
    """Cache clear endpoint."""
    classifier = current_app.classifier
    if classifier is None:
        g.response.set_data({"message": "キャッシュをクリアできません（分類器が初期化されていません）"})
        return jsonify(g.response.get_result())

    try:
        classifier.clear_cache()
        g.response.set_data({"message": "キャッシュをクリアしました"})
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("キャッシュクリアエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train/data", methods=["GET"])
def get_training_data():
    """Get paginated training data."""
    global _training_data_store
    try:
        page = request.args.get('page', 1, type=int)
        page_size = request.args.get('page_size', 20, type=int)
        page = max(1, page)
        page_size = max(1, min(100, page_size))

        total = len(_training_data_store)
        total_pages = max(1, (total + page_size - 1) // page_size)
        start = (page - 1) * page_size
        end = start + page_size
        items = _training_data_store[start:end]

        g.response.set_data({
            "items": items,
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages
        })
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("学習データ取得エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train/data", methods=["POST"])
def upload_training_data():
    """Upload training data (JSON array of {text, label})."""
    global _training_data_store
    try:
        data = request.get_json()
        if not data or 'training_data' not in data:
            g.response.add_error_message("「training_data」フィールドは必須です（{text, label} の配列）")
            return jsonify(g.response.get_result()), 422

        training_data = data['training_data']
        if not isinstance(training_data, list):
            g.response.add_error_message("「training_data」は配列である必要があります")
            return jsonify(g.response.get_result()), 422

        mode = data.get('mode', 'replace')  # 'replace' or 'append'
        valid_items, reject_stats = _sanitize_training_items(training_data)

        if len(valid_items) == 0:
            g.response.add_error_message("有効なデータがありません。各要素には「text」と「label」が必要です。")
            return jsonify(g.response.get_result()), 422

        if mode == 'append':
            _training_data_store.extend(valid_items)
        else:
            _training_data_store = valid_items

        g.response.set_data({
            "message": "%d 件をアップロードしました（mode: %s）" % (len(valid_items), mode),
            "total": len(_training_data_store),
            "rejected": reject_stats
        })
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("学習データアップロードエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train/data/upload", methods=["POST"])
def upload_training_data_xlsx():
    """Upload training data from xlsx or csv file (text, label columns).
    
    The uploaded file will be saved as 'training_data.xlsx' or 'training_data.csv' (overwrite existing).
    Expects file with 2 columns: 'text' and 'label' (or first 2 columns if no headers match).
    Supported formats: .xlsx, .csv
    """
    global _training_data_store
    try:
        # ファイル有無を確認
        if 'file' not in request.files:
            g.response.add_error_message("ファイルがありません。.xlsx または .csv ファイルをアップロードしてください。")
            return jsonify(g.response.get_result()), 422
        
        file = request.files['file']
        if file.filename == '':
            g.response.add_error_message("ファイルが選択されていません。")
            return jsonify(g.response.get_result()), 422
        
        # 拡張子を検証
        filename_lower = file.filename.lower()
        is_xlsx = filename_lower.endswith('.xlsx')
        is_csv = filename_lower.endswith('.csv')
        
        if not is_xlsx and not is_csv:
            g.response.add_error_message("ファイル形式が不正です。.xlsx / .csv を使用してください。")
            return jsonify(g.response.get_result()), 422
        
        # データディレクトリを作成
        data_dir = os.path.dirname(_TRAINING_DATA_FILE)
        os.makedirs(data_dir, exist_ok=True)
        
        valid_items = []
        skipped = 0
        saved_filename = ""
        
        if is_xlsx:
            # openpyxl を読み込み
            try:
                from openpyxl import load_workbook
            except ImportError:
                g.response.add_error_message("openpyxl がインストールされていません。pip install openpyxl でインストールしてください。")
                return jsonify(g.response.get_result()), 500
            
            # training_data.xlsx として保存（既存は上書き）
            xlsx_path = os.path.join(data_dir, "training_data.xlsx")
            file.save(xlsx_path)
            saved_filename = "training_data.xlsx"
            
            # xlsx を読み込み解析
            wb = load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            
            # 全行を取得
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                g.response.add_error_message("ファイルが空、またはデータ行がありません。")
                return jsonify(g.response.get_result()), 422
            
            # ヘッダ行を検出（先頭行に text/label があるか）
            first_row = rows[0]
            text_col_idx = None
            label_col_idx = None
            has_header = False
            
            # ヘッダ名から列を推定（大文字小文字無視）
            if first_row and len(first_row) >= 2:
                header_lower = [str(h).lower().strip() if h else '' for h in first_row]
                if 'text' in header_lower:
                    text_col_idx = header_lower.index('text')
                if 'label' in header_lower:
                    label_col_idx = header_lower.index('label')
                
                if text_col_idx is not None and label_col_idx is not None:
                    has_header = True
            
            # ヘッダが見つからない場合は先頭 2 列を使用（text=0, label=1）
            if text_col_idx is None:
                text_col_idx = 0
            if label_col_idx is None:
                label_col_idx = 1
            
            # データ行を解析
            data_rows = rows[1:] if has_header else rows
            
            for row in data_rows:
                if row and len(row) > max(text_col_idx, label_col_idx):
                    text_val = row[text_col_idx]
                    label_val = row[label_col_idx]
                    
                    if text_val is not None and label_val is not None:
                        text_str = str(text_val).strip()
                        label_str = str(label_val).strip()
                        
                        if text_str and label_str:
                            valid_items.append({'text': text_str, 'label': label_str})
                        else:
                            skipped += 1
                    else:
                        skipped += 1
                else:
                    skipped += 1
            
            wb.close()
            
        else:  # CSV
            import csv
            import io
            
            # 先にコンテンツを読み込む（file.save()後はストリームが消費されるため）
            content = file.stream.read().decode('utf-8-sig')  # BOM対応
            file.stream.seek(0)  # リセットしてから保存
            
            # training_data.csv として保存（既存は上書き）
            csv_path = os.path.join(data_dir, "training_data.csv")
            file.save(csv_path)
            saved_filename = "training_data.csv"
            
            # CSVを解析
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            
            if len(rows) < 2:
                g.response.add_error_message("ファイルが空、またはデータ行がありません。")
                return jsonify(g.response.get_result()), 422
            
            # ヘッダ行を検出
            first_row = rows[0]
            text_col_idx = None
            label_col_idx = None
            has_header = False
            
            if first_row and len(first_row) >= 2:
                header_lower = [str(h).lower().strip() if h else '' for h in first_row]
                if 'text' in header_lower:
                    text_col_idx = header_lower.index('text')
                if 'label' in header_lower:
                    label_col_idx = header_lower.index('label')
                
                if text_col_idx is not None and label_col_idx is not None:
                    has_header = True
            
            # ヘッダが見つからない場合は先頭 2 列を使用
            if text_col_idx is None:
                text_col_idx = 0
            if label_col_idx is None:
                label_col_idx = 1
            
            # データ行を解析
            data_rows = rows[1:] if has_header else rows
            
            for row in data_rows:
                if row and len(row) > max(text_col_idx, label_col_idx):
                    text_val = row[text_col_idx]
                    label_val = row[label_col_idx]
                    
                    if text_val is not None and label_val is not None:
                        text_str = str(text_val).strip()
                        label_str = str(label_val).strip()
                        
                        if text_str and label_str:
                            valid_items.append({'text': text_str, 'label': label_str})
                        else:
                            skipped += 1
                    else:
                        skipped += 1
                else:
                    skipped += 1
        
        if len(valid_items) == 0:
            g.response.add_error_message("有効なデータがありません。各行に「text」と「label」の値が必要です。")
            return jsonify(g.response.get_result()), 422
        
        # 学習データストアを置換（上書き）
        _training_data_store = valid_items
        profile = _build_training_profile(_training_data_store)

        file_type = "xlsx" if is_xlsx else "csv"
        g.response.set_data({
            "message": "%s から学習サンプル %d 件をアップロードしました" % (file_type, len(valid_items)),
            "total": len(_training_data_store),
            "file_saved": saved_filename,
            "skipped_rows": skipped,
            "profile": profile
        })
        return jsonify(g.response.get_result())
        
    except Exception as e:
        logging.error("学習データアップロードエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train/data", methods=["DELETE"])
def clear_training_data():
    """Clear all training data."""
    global _training_data_store
    _training_data_store = []
    g.response.set_data({"message": "学習データをクリアしました", "total": 0})
    return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/train/data/download", methods=["GET"])
def download_training_data():
    """Download all training data as JSON file."""
    global _training_data_store
    response = make_response(json.dumps(_training_data_store, ensure_ascii=False, indent=2))
    response.headers['Content-Type'] = 'application/json'
    response.headers['Content-Disposition'] = 'attachment; filename=training_data.json'
    return response


@api_blueprint.route("/api/v1/train/profile", methods=["GET", "POST"])
def profile_training_data():
    """Profile training data quality and return recommendations."""
    global _training_data_store
    try:
        if request.method == "POST":
            data = request.get_json(silent=True) or {}
            raw_training_data = data.get("training_data", None)
            if raw_training_data is None:
                g.response.add_error_message("「training_data」フィールドは必須です")
                return jsonify(g.response.get_result()), 422
            valid_items, reject_stats = _sanitize_training_items(raw_training_data)
        else:
            valid_items = list(_training_data_store)
            reject_stats = {
                "invalid_item_count": 0,
                "missing_field_count": 0,
                "empty_text_count": 0,
                "empty_label_count": 0
            }

        profile = _build_training_profile(valid_items)
        profile["rejected"] = reject_stats
        profile["source"] = "request_payload" if request.method == "POST" else "in_memory_store"
        g.response.set_data(profile)
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("学習データ診断エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train/validate", methods=["POST"])
def validate_training_data():
    """Validate training data quality without training."""
    try:
        data = request.get_json()
        if not data or 'training_data' not in data:
            g.response.add_error_message("「training_data」フィールドは必須です（{text, label} の配列）")
            return jsonify(g.response.get_result()), 422

        training_data = data['training_data']
        if not isinstance(training_data, list) or len(training_data) == 0:
            g.response.add_error_message("「training_data」は空でない配列である必要があります")
            return jsonify(g.response.get_result()), 422

        valid_items, reject_stats = _sanitize_training_items(training_data)
        if len(valid_items) == 0:
            g.response.add_error_message("有効なデータがありません。各要素の「text」と「label」は空でない必要があります。")
            return jsonify(g.response.get_result()), 422

        profile = _build_training_profile(valid_items)
        issue_levels = {
            "errors": [i["message"] for i in profile["issue_details"] if i["level"] == "error"],
            "warnings": [i["message"] for i in profile["issue_details"] if i["level"] == "warning"],
            "info": [i["message"] for i in profile["issue_details"] if i["level"] == "info"]
        }
        g.response.set_data({
            "valid": profile["quality_gate_passed"],
            "total_samples": profile["total_samples"],
            "num_classes": profile["num_classes"],
            "class_distribution": profile["class_distribution"],
            "issues": profile["issues"],
            "issue_levels": issue_levels,
            "health_score": profile["health_score"],
            "readiness": profile["readiness"],
            "imbalance_ratio": profile["imbalance_ratio"],
            "duplicate_ratio": profile["duplicate_ratio"],
            "text_length_stats": profile["text_length_stats"],
            "recommendations": profile["recommendations"],
            "suggested_params": profile["suggested_params"],
            "rejected": reject_stats
        })
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("バリデーションエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train", methods=["POST"])
def start_training():
    """Start model training (async background task)."""
    global _training_state
    global _training_data_store

    with _training_state_lock:
        if _training_state['status'] == 'running':
            g.response.add_error_message("学習はすでに実行中です")
            return jsonify(g.response.get_result()), 409

    app = current_app._get_current_object()
    previous_model_summary = _extract_previous_model_summary(getattr(app, "classifier", None))

    try:
        data = request.get_json() or {}

        training_data = data.get('training_data', None)
        # If no training_data provided in request, use the in-memory store
        if not training_data:
            if len(_training_data_store) == 0:
                g.response.add_error_message("学習データがありません。先にアップロードするか、リクエストで「training_data」を指定してください。")
                return jsonify(g.response.get_result()), 422
            training_data = _training_data_store
        if not isinstance(training_data, list) or len(training_data) == 0:
            g.response.add_error_message("「training_data」は空でない配列である必要があります")
            return jsonify(g.response.get_result()), 422

        valid_items, reject_stats = _sanitize_training_items(training_data)
        if len(valid_items) == 0:
            g.response.add_error_message("有効なデータがありません。各要素の「text」と「label」は空でない必要があります。")
            return jsonify(g.response.get_result()), 422

        profile = _build_training_profile(valid_items)
        if not profile["quality_gate_passed"]:
            g.response.add_error_message(
                "学習データが品質ゲートを通過していません: %s" %
                ("; ".join(profile["issues"][:2]) if profile["issues"] else "品質不足")
            )
            g.response.set_data({
                "profile": profile,
                "rejected": reject_stats
            })
            return jsonify(g.response.get_result()), 422

        texts = [item["text"] for item in valid_items]
        labels = [item["label"] for item in valid_items]

        # Training parameters with profile-aware defaults
        params = data.get('params', {}) if isinstance(data.get('params', {}), dict) else {}
        suggested = profile.get("suggested_params", {})
        test_size = params.get('test_size', suggested.get("test_size", 0.15))
        n_estimators = params.get('n_estimators', suggested.get("n_estimators", 300))
        learning_rate = params.get('learning_rate', suggested.get("learning_rate", 0.05))
        max_depth = params.get('max_depth', suggested.get("max_depth", 6))
        algorithm_strategy = str(params.get('algorithm_strategy', suggested.get("algorithm_strategy", 'auto'))).strip().lower()
        compare_baselines = _to_bool(
            params.get('compare_baselines', suggested.get("compare_baselines", True)),
            default=True
        )
        auto_tune = _to_bool(
            params.get('auto_tune', suggested.get("auto_tune", True)),
            default=True
        )
        rebalance_strategy = str(
            params.get('rebalance_strategy', suggested.get("rebalance_strategy", "none"))
        ).strip().lower()
        try:
            test_size = float(test_size)
            n_estimators = int(n_estimators)
            learning_rate = float(learning_rate)
            max_depth = int(max_depth)
            random_state = int(params.get('random_state', 42))
        except (TypeError, ValueError):
            g.response.add_error_message("学習パラメータの型が不正です。数値項目を確認してください。")
            return jsonify(g.response.get_result()), 422

        if not (0.05 <= test_size <= 0.40):
            g.response.add_error_message("「test_size」は 0.05〜0.40 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422
        if not (50 <= n_estimators <= 1000):
            g.response.add_error_message("「n_estimators」は 50〜1000 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422
        if not (0.01 <= learning_rate <= 0.30):
            g.response.add_error_message("「learning_rate」は 0.01〜0.30 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422
        if not (2 <= max_depth <= 10):
            g.response.add_error_message("「max_depth」は 2〜10 の範囲で指定してください")
            return jsonify(g.response.get_result()), 422
        if algorithm_strategy not in ('auto', 'gbdt', 'lr'):
            g.response.add_error_message("「algorithm_strategy」は auto / gbdt / lr のいずれかを指定してください")
            return jsonify(g.response.get_result()), 422
        if rebalance_strategy not in ('none', 'balanced_upsample', 'auto'):
            g.response.add_error_message("「rebalance_strategy」は none / balanced_upsample / auto のいずれかを指定してください")
            return jsonify(g.response.get_result()), 422
        if random_state < 0:
            g.response.add_error_message("「random_state」は 0 以上で指定してください")
            return jsonify(g.response.get_result()), 422
        if algorithm_strategy in ('gbdt', 'lr'):
            compare_baselines = False

        effective_rebalance = (
            'balanced_upsample'
            if rebalance_strategy == 'auto' and profile["imbalance_ratio"] >= 3
            else ('none' if rebalance_strategy == 'auto' else rebalance_strategy)
        )

        effective_params = {
            "test_size": test_size,
            "n_estimators": n_estimators,
            "learning_rate": learning_rate,
            "max_depth": max_depth,
            "algorithm_strategy": algorithm_strategy,
            "compare_baselines": bool(compare_baselines),
            "auto_tune": bool(auto_tune),
            "rebalance_strategy": effective_rebalance,
            "random_state": random_state
        }

        # Reset training state
        with _training_state_lock:
            _training_state = {
                'status': 'running',
                'progress': '初期化中…',
                'results': None,
                'error': None,
                'started_at': dt.datetime.now().isoformat(),
                'finished_at': None,
                'dataset_profile': profile,
                'params': effective_params,
                'previous_model_summary': previous_model_summary
            }

        def run_training():
            global _training_state
            try:
                from denpyo_toroku.src.denpyo_toroku.classifier import ProductionIntentClassifier

                with _training_state_lock:
                    _training_state['progress'] = '分類器を初期化中…'

                def progress_callback(message: str):
                    with _training_state_lock:
                        _training_state['progress'] = str(message)

                classifier = ProductionIntentClassifier(
                    config_path=AppConfig.OCI_CONFIG_PATH,
                    profile=AppConfig.OCI_CONFIG_PROFILE,
                    service_endpoint=AppConfig.OCI_SERVICE_ENDPOINT,
                    compartment_id=AppConfig.OCI_CONFIG_COMPARTMENT,
                    embedding_model_id=AppConfig.EMBEDDING_MODEL_ID,
                    log_level=AppConfig.LOG_LEVEL,
                    enable_cache=False,  # 学習中はキャッシュしない
                    enable_monitoring=True
                )

                with _training_state_lock:
                    _training_state['progress'] = 'モデルを学習中（数分かかる場合があります）…'

                results = classifier.train(
                    texts=texts,
                    labels=labels,
                    test_size=test_size,
                    random_state=random_state,
                    validation_split=0.15,
                    early_stopping_rounds=15,
                    compare_baselines=bool(compare_baselines),
                    preferred_algorithm=algorithm_strategy,
                    auto_tune=bool(auto_tune),
                    rebalance_strategy=effective_rebalance,
                    progress_callback=progress_callback,
                    n_estimators=n_estimators,
                    learning_rate=learning_rate,
                    max_depth=max_depth,
                    min_samples_split=15,
                    min_samples_leaf=8,
                    subsample=0.8,
                    n_iter_no_change=15,
                    tol=1e-4
                )

                # 品質評価
                quality_ok = True
                quality_issues = []
                if results['test_accuracy'] < 0.85:
                    quality_issues.append('テスト精度が 85% を下回っています: %.2f%%' % (results['test_accuracy'] * 100))
                    quality_ok = False
                if results.get('test_macro_f1', 0.0) < 0.80:
                    quality_issues.append('Macro-F1 が 0.80 を下回っています: %.4f' % results.get('test_macro_f1', 0.0))
                    quality_ok = False
                if results['overfitting_gap'] > 0.10:
                    quality_issues.append('過学習の可能性: ギャップ %.4f > 0.10' % results['overfitting_gap'])
                    quality_ok = False
                if results.get('macro_f1_gap', 0.0) > 0.12:
                    quality_issues.append('Macro-F1 ギャップが大きく過学習リスクがあります: %.4f > 0.12' % results.get('macro_f1_gap', 0.0))
                    quality_ok = False
                if profile["readiness"] == "low":
                    quality_issues.append('データ準備度が低いです。本番展開前に診断結果の問題修正を検討してください。')

                # モデルを保存（features/train_production.py と同様にバックアップ作成）
                with _training_state_lock:
                    _training_state['progress'] = 'モデルを保存中…'
                model_path = AppConfig.MODEL_PATH
                if os.path.exists(model_path):
                    timestamp = dt.datetime.now().strftime('%Y%m%d_%H%M%S')
                    backup_path = "%s.backup_%s" % (model_path, timestamp)
                    os.rename(model_path, backup_path)

                classifier.save_model(model_path, include_metadata=True)

                # 実行中アプリへモデルを反映
                with _training_state_lock:
                    _training_state['progress'] = 'サービスにモデルを反映中…'
                app.classifier.load_model(model_path) if app.classifier else None
                if app.classifier is None:
                    app.classifier = classifier

                with _training_state_lock:
                    _training_state['status'] = 'completed'
                    _training_state['progress'] = '学習が完了しました'
                    comparison_with_previous = _build_model_comparison(previous_model_summary, results)
                    _training_state['results'] = {
                        'train_accuracy': results['train_accuracy'],
                        'test_accuracy': results['test_accuracy'],
                        'overfitting_gap': results['overfitting_gap'],
                        'train_macro_f1': results.get('train_macro_f1', 0.0),
                        'test_macro_f1': results.get('test_macro_f1', 0.0),
                        'test_weighted_f1': results.get('test_weighted_f1', 0.0),
                        'macro_f1_gap': results.get('macro_f1_gap', 0.0),
                        'selection_score': results.get('selection_score', 0.0),
                        'selected_algorithm': results.get('selected_algorithm', 'GradientBoostingClassifier'),
                        'requested_algorithm': results.get('requested_algorithm', algorithm_strategy),
                        'params_used': effective_params,
                        'candidates': results.get('candidates', []),
                        'per_class_metrics': results.get('per_class_metrics', []),
                        'num_classes': results['num_classes'],
                        'train_samples': results['train_samples'],
                        'test_samples': results['test_samples'],
                        'training_duration_seconds': results.get('training_duration_seconds', 0.0),
                        'n_estimators_used': results['n_estimators_used'],
                        'dataset_profile': profile,
                        'recommendations': profile.get("recommendations", []),
                        'previous_model_summary': previous_model_summary,
                        'comparison_with_previous': comparison_with_previous,
                        'quality_ok': quality_ok,
                        'quality_issues': quality_issues,
                        'model_path': model_path
                    }
                    _training_state['finished_at'] = dt.datetime.now().isoformat()

            except Exception as e:
                logging.error("学習失敗: %s", e, exc_info=True)
                with _training_state_lock:
                    _training_state['status'] = 'failed'
                    _training_state['error'] = str(e)
                    _training_state['progress'] = '学習に失敗しました'
                    _training_state['finished_at'] = dt.datetime.now().isoformat()

        thread = threading.Thread(target=run_training, daemon=True)
        thread.start()

        g.response.set_data({
            "message": "学習を開始しました",
            "status": "running",
            "profile": profile,
            "params": effective_params
        })
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("学習開始エラー: %s", e, exc_info=True)
        with _training_state_lock:
            _training_state['status'] = 'idle'
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/train/status", methods=["GET"])
def get_training_status():
    """Get current training status."""
    with _training_state_lock:
        g.response.set_data(dict(_training_state))
    return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/model/reload", methods=["POST"])
def reload_model():
    """Reload model from disk."""
    try:
        model_path = AppConfig.MODEL_PATH
        if not os.path.exists(model_path):
            g.response.add_error_message("モデルファイルが見つかりません: %s" % model_path)
            return jsonify(g.response.get_result()), 404

        classifier = current_app.classifier
        if classifier is None:
            from denpyo_toroku.src.denpyo_toroku.classifier import ProductionIntentClassifier
            classifier = ProductionIntentClassifier(
                config_path=AppConfig.OCI_CONFIG_PATH,
                profile=AppConfig.OCI_CONFIG_PROFILE,
                service_endpoint=AppConfig.OCI_SERVICE_ENDPOINT,
                compartment_id=AppConfig.OCI_CONFIG_COMPARTMENT,
                embedding_model_id=AppConfig.EMBEDDING_MODEL_ID,
                log_level=AppConfig.LOG_LEVEL,
                enable_cache=AppConfig.ENABLE_CACHE,
                cache_size=AppConfig.CACHE_SIZE,
                enable_monitoring=True
            )
            current_app.classifier = classifier

        classifier.load_model(model_path)
        g.response.set_data({"message": "モデルを再読み込みしました: %s" % model_path})
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("モデル再読み込みエラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/model/info", methods=["GET"])
def get_model_info():
    """Model information endpoint."""
    classifier = current_app.classifier
    if classifier is None or classifier.classifier is None:
        g.response.set_data({
            "classes": [],
            "num_classes": 0,
            "embedding_model": "-",
            "algorithm": "GradientBoostingClassifier",
            "n_estimators": 0,
            "embedding_dimension": 0,
            "model_loaded": False,
            "message": "分類器が初期化されていません" if classifier is None else "モデルが読み込まれていません"
        })
        return jsonify(g.response.get_result())

    try:
        info = {
            "classes": list(classifier.label_encoder.keys()),
            "num_classes": len(classifier.label_encoder),
            "embedding_model": classifier.embedding_model_id,
            "algorithm": getattr(classifier, "algorithm_name", "GradientBoostingClassifier"),
            "n_estimators": getattr(classifier.classifier, "n_estimators_", 0),
            "embedding_dimension": classifier.classifier.n_features_in_ if hasattr(classifier.classifier, 'n_features_in_') else 0,
            "model_source": getattr(classifier, "model_source", "unknown"),
            "model_timestamp": getattr(classifier, "model_timestamp", None),
            "training_summary": getattr(classifier, "last_training_summary", None)
        }
        g.response.set_data(info)
        return jsonify(g.response.get_result())
    except Exception as e:
        logging.error("モデル情報取得エラー: %s", e, exc_info=True)
        g.response.add_error_message(str(e))
        return jsonify(g.response.get_result()), 500


# ========================================
# ADB (Autonomous Database) Management API
# ========================================

def _get_oci_config_for_adb() -> Optional[Dict[str, Any]]:
    """OCI設定を取得してADB操作用のconfigを返す

    ページの「リージョン」フィールドに表示される値（.envのOCI_REGION）と
    OCIクライアントが使用するregionを統一するため、~/.oci/configのregionを
    .envのOCI_REGIONで上書きする。
    """
    try:
        config_path = os.path.expanduser(AppConfig.OCI_CONFIG_PATH)
        profile = AppConfig.OCI_CONFIG_PROFILE

        if not os.path.exists(config_path):
            logging.warning("OCI設定ファイルが見つかりません: %s", config_path)
            return None

        import oci
        config = oci.config.from_file(config_path, profile)

        # ページのリージョンフィールドと統一するため、.envのOCI_REGIONで上書き
        # これにより「再取得」時に使用されるregionはページに表示される値と同じになる
        env_region = os.environ.get("OCI_REGION", AppConfig.OCI_REGION)
        if env_region:
            config["region"] = env_region

        return config
    except Exception as e:
        logging.error("OCI設定読み込みエラー: %s", e, exc_info=True)
        return None


def _get_adb_client():
    """ADBクライアントを取得"""
    try:
        import oci
        config = _get_oci_config_for_adb()
        if not config:
            return None
        return oci.database.DatabaseClient(config)
    except Exception as e:
        logging.error("ADBクライアント作成エラー: %s", e, exc_info=True)
        return None


@api_blueprint.route("/api/v1/database/adb/info", methods=["GET"])
def get_adb_info():
    """Autonomous Database情報を取得"""
    adb_ocid = os.environ.get(_DB_ADB_OCID_ENV_KEY, "").strip()

    # .envからregionを取得（デフォルトはAppConfig.OCI_REGION = ap-osaka-1）
    region = os.environ.get("OCI_REGION", AppConfig.OCI_REGION)

    if not adb_ocid:
        g.response.set_data({
            "status": "not_configured",
            "message": "ADB OCID が設定されていません。",
            "id": None,
            "display_name": None,
            "lifecycle_state": None,
            "region": region
        })
        return jsonify(g.response.get_result())

    try:
        db_client = _get_adb_client()
        if not db_client:
            g.response.set_data({
                "status": "error",
                "message": "OCI接続を確認できません。OCI設定を確認してください。",
                "id": adb_ocid,
                "display_name": None,
                "lifecycle_state": None,
                "region": region,
            })
            return jsonify(g.response.get_result())

        adb = db_client.get_autonomous_database(adb_ocid).data

        g.response.set_data({
            "status": "success",
            "message": "データベース情報を取得しました。",
            "id": adb.id,
            "display_name": adb.display_name,
            "lifecycle_state": adb.lifecycle_state,
            "db_name": getattr(adb, "db_name", None),
            "cpu_core_count": getattr(adb, "cpu_core_count", None),
            "data_storage_size_in_tbs": getattr(adb, "data_storage_size_in_tbs", None),
            "region": region
        })
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("ADB情報取得エラー: %s", e, exc_info=True)
        g.response.set_data({
            "status": "error",
            "message": f"データベース情報の取得に失敗しました: {str(e)}",
            "id": adb_ocid,
            "display_name": None,
            "lifecycle_state": None,
            "region": region,
        })
        return jsonify(g.response.get_result())


@api_blueprint.route("/api/v1/database/adb/start", methods=["POST"])
def start_adb():
    """Autonomous Databaseを起動"""
    adb_ocid = os.environ.get(_DB_ADB_OCID_ENV_KEY, "").strip()
    region = os.environ.get("OCI_REGION", AppConfig.OCI_REGION)

    if not adb_ocid:
        g.response.add_error_message("ADB OCID が設定されていません。")
        return jsonify(g.response.get_result()), 400

    try:
        db_client = _get_adb_client()
        if not db_client:
            g.response.add_error_message("OCI接続を確認できません。OCI設定を確認してください。")
            return jsonify(g.response.get_result()), 500

        # 現在の状態を確認
        adb = db_client.get_autonomous_database(adb_ocid).data

        if adb.lifecycle_state == "AVAILABLE":
            g.response.set_data({
                "status": "already_available",
                "message": "データベースは既に起動しています。",
                "id": adb.id,
                "display_name": adb.display_name,
                "lifecycle_state": adb.lifecycle_state,
                "region": region,
            })
            return jsonify(g.response.get_result())

        if adb.lifecycle_state not in ["STOPPED", "UNAVAILABLE"]:
            g.response.set_data({
                "status": "cannot_start",
                "message": f"データベースの現在の状態 ({adb.lifecycle_state}) では起動できません。",
                "id": adb.id,
                "display_name": adb.display_name,
                "lifecycle_state": adb.lifecycle_state,
                "region": region,
            })
            return jsonify(g.response.get_result())

        # 起動リクエスト送信
        db_client.start_autonomous_database(adb_ocid)

        g.response.set_data({
            "status": "accepted",
            "message": f"データベース '{adb.display_name}' の起動を開始しました。",
            "id": adb.id,
            "display_name": adb.display_name,
            "lifecycle_state": "STARTING",
            "region": region,
        })
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("ADB起動エラー: %s", e, exc_info=True)
        g.response.add_error_message(f"データベースの起動に失敗しました: {str(e)}")
        return jsonify(g.response.get_result()), 500


@api_blueprint.route("/api/v1/database/adb/stop", methods=["POST"])
def stop_adb():
    """Autonomous Databaseを停止"""
    adb_ocid = os.environ.get(_DB_ADB_OCID_ENV_KEY, "").strip()
    region = os.environ.get("OCI_REGION", AppConfig.OCI_REGION)

    if not adb_ocid:
        g.response.add_error_message("ADB OCID が設定されていません。")
        return jsonify(g.response.get_result()), 400

    try:
        db_client = _get_adb_client()
        if not db_client:
            g.response.add_error_message("OCI接続を確認できません。OCI設定を確認してください。")
            return jsonify(g.response.get_result()), 500

        # 現在の状態を確認
        adb = db_client.get_autonomous_database(adb_ocid).data

        if adb.lifecycle_state == "STOPPED":
            g.response.set_data({
                "status": "already_stopped",
                "message": "データベースは既に停止しています。",
                "id": adb.id,
                "display_name": adb.display_name,
                "lifecycle_state": adb.lifecycle_state,
                "region": region,
            })
            return jsonify(g.response.get_result())

        if adb.lifecycle_state not in ["AVAILABLE"]:
            g.response.set_data({
                "status": "cannot_stop",
                "message": f"データベースの現在の状態 ({adb.lifecycle_state}) では停止できません。",
                "id": adb.id,
                "display_name": adb.display_name,
                "lifecycle_state": adb.lifecycle_state,
                "region": region,
            })
            return jsonify(g.response.get_result())

        # 停止リクエスト送信
        db_client.stop_autonomous_database(adb_ocid)

        g.response.set_data({
            "status": "accepted",
            "message": f"データベース '{adb.display_name}' の停止を開始しました。",
            "id": adb.id,
            "display_name": adb.display_name,
            "lifecycle_state": "STOPPING",
            "region": region,
        })
        return jsonify(g.response.get_result())

    except Exception as e:
        logging.error("ADB停止エラー: %s", e, exc_info=True)
        g.response.add_error_message(f"データベースの停止に失敗しました: {str(e)}")
        return jsonify(g.response.get_result()), 500
